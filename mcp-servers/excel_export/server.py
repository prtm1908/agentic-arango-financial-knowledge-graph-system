#!/usr/bin/env python3
"""
Excel Export MCP Server

Provides tools for:
- Creating Excel workbooks from financial metrics
- Formatting financial reports
- Generating comparison spreadsheets
"""

import os
import json
import time
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent
try:
    import redis
except Exception:  # pragma: no cover - optional dependency
    redis = None

# Configuration
OUTPUT_PATH = os.getenv("OUTPUT_PATH", "/output")
os.makedirs(OUTPUT_PATH, exist_ok=True)
REDIS_URL = os.getenv("REDIS_URL", "")
OPENCODE_JOB_ID = os.getenv("OPENCODE_JOB_ID", "")
OPENCODE_AGENT_NAME = os.getenv("OPENCODE_AGENT_NAME", "")

# Create MCP server
server = Server("excel_export")
_redis_client = None


def _get_redis_client():
    global _redis_client
    if _redis_client is not None:
        return _redis_client
    if not REDIS_URL or redis is None:
        return None
    try:
        _redis_client = redis.from_url(REDIS_URL, decode_responses=True)
    except Exception:
        _redis_client = None
    return _redis_client


def _publish_event(event: dict):
    if not OPENCODE_JOB_ID:
        return
    client = _get_redis_client()
    if client is None:
        return
    payload = dict(event)
    payload.setdefault("timestamp", int(time.time() * 1000))
    event_json = json.dumps(payload)
    history_key = f"event_history:{OPENCODE_JOB_ID}"
    channel = f"events:{OPENCODE_JOB_ID}"
    try:
        client.rpush(history_key, event_json)
        client.ltrim(history_key, -100, -1)
        client.expire(history_key, 300)
        client.publish(channel, event_json)
    except Exception:
        return


def _publish_tool_call(name: str, args: dict):
    payload = {"type": "tool_call", "tool": name, "server": "mcp", "args": args}
    if OPENCODE_AGENT_NAME:
        payload["agent"] = OPENCODE_AGENT_NAME
    _publish_event(payload)


def _publish_tool_result(name: str, result: dict, duration_ms: int):
    payload = {"type": "tool_result", "tool": name, "result": result, "duration_ms": duration_ms}
    if OPENCODE_AGENT_NAME:
        payload["agent"] = OPENCODE_AGENT_NAME
    _publish_event(payload)


def _tool_response(name: str, payload: dict, started_at: float) -> list[TextContent]:
    duration_ms = int((time.time() - started_at) * 1000)
    _publish_tool_result(name, payload, duration_ms)
    return [TextContent(type="text", text=json.dumps(payload))]

# Styles
HEADER_FONT = Font(bold=True, size=12)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=12, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def format_worksheet(ws, headers: list[str]):
    """Apply standard formatting to a worksheet."""
    # Set header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(str(headers[col-1]))
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)


@server.list_tools()
async def list_tools() -> list[Tool]:
    """List available tools."""
    return [
        Tool(
            name="create_metrics_report",
            description="Create an Excel report from extracted metrics",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "Output filename (without .xlsx extension)"
                    },
                    "title": {
                        "type": "string",
                        "description": "Report title"
                    },
                    "company_name": {
                        "type": "string",
                        "description": "Company name"
                    },
                    "metrics": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "metric_name": {"type": "string"},
                                "value": {"type": "number"},
                                "unit": {"type": "string"},
                                "denomination": {"type": "string"},
                                "fiscal_year": {"type": "string"},
                                "source_pages": {"type": "array", "items": {"type": "integer"}}
                            }
                        },
                        "description": "Array of metric objects"
                    }
                },
                "required": ["filename", "title", "company_name", "metrics"]
            }
        ),
        Tool(
            name="create_comparison_report",
            description="Create a multi-company or multi-period comparison report",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "Output filename"
                    },
                    "title": {
                        "type": "string",
                        "description": "Report title"
                    },
                    "comparison_data": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "entity": {"type": "string"},
                                "metrics": {"type": "object"}
                            }
                        },
                        "description": "Array of entities with their metrics"
                    },
                    "metric_names": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of metrics to include in comparison"
                    }
                },
                "required": ["filename", "title", "comparison_data", "metric_names"]
            }
        ),
        Tool(
            name="create_time_series_report",
            description="Create a time series report for metrics across periods",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "Output filename"
                    },
                    "title": {
                        "type": "string",
                        "description": "Report title"
                    },
                    "company_name": {
                        "type": "string",
                        "description": "Company name"
                    },
                    "periods": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of periods (e.g., ['FY22', 'FY23', 'FY24'])"
                    },
                    "metrics_by_period": {
                        "type": "object",
                        "description": "Object with periods as keys and metric arrays as values"
                    }
                },
                "required": ["filename", "title", "company_name", "periods", "metrics_by_period"]
            }
        )
    ]


@server.call_tool()
async def call_tool(name: str, arguments: dict[str, Any]) -> list[TextContent]:
    """Handle tool calls."""
    started_at = time.time()
    _publish_tool_call(name, arguments)

    if name == "create_metrics_report":
        filename = arguments["filename"]
        title = arguments["title"]
        company_name = arguments["company_name"]
        metrics = arguments["metrics"]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Metrics"

            # Title row
            ws.merge_cells('A1:F1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Company info
            ws['A2'] = f"Company: {company_name}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

            # Headers
            headers = ["Metric Name", "Value", "Unit", "Denomination", "Fiscal Year", "Source Pages"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = HEADER_FONT_WHITE
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER

            # Data rows
            for row_idx, metric in enumerate(metrics, 6):
                ws.cell(row=row_idx, column=1, value=metric.get("metric_name", "")).border = THIN_BORDER
                ws.cell(row=row_idx, column=2, value=metric.get("value")).border = THIN_BORDER
                ws.cell(row=row_idx, column=3, value=metric.get("unit", "")).border = THIN_BORDER
                ws.cell(row=row_idx, column=4, value=metric.get("denomination", "")).border = THIN_BORDER
                ws.cell(row=row_idx, column=5, value=metric.get("fiscal_year", "")).border = THIN_BORDER

                pages = metric.get("source_pages", [])
                pages_str = ", ".join(str(p) for p in pages) if pages else ""
                ws.cell(row=row_idx, column=6, value=pages_str).border = THIN_BORDER

            # Auto-adjust columns
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 18

            # Save
            output_file = os.path.join(OUTPUT_PATH, f"{filename}.xlsx")
            wb.save(output_file)

            payload = {
                "success": True,
                "file_path": output_file,
                "metrics_count": len(metrics)
            }
            return _tool_response(name, payload, started_at)

        except Exception as e:
            payload = {"error": str(e)}
            return _tool_response(name, payload, started_at)

    elif name == "create_comparison_report":
        filename = arguments["filename"]
        title = arguments["title"]
        comparison_data = arguments["comparison_data"]
        metric_names = arguments["metric_names"]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison"

            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(metric_names)+1)
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=16)

            # Headers: Entity | Metric1 | Metric2 | ...
            headers = ["Entity"] + metric_names
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = HEADER_FONT_WHITE
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER

            # Data rows
            for row_idx, entity_data in enumerate(comparison_data, 4):
                ws.cell(row=row_idx, column=1, value=entity_data.get("entity", "")).border = THIN_BORDER

                metrics = entity_data.get("metrics", {})
                for col_idx, metric_name in enumerate(metric_names, 2):
                    value = metrics.get(metric_name, "")
                    ws.cell(row=row_idx, column=col_idx, value=value).border = THIN_BORDER

            # Auto-adjust
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            output_file = os.path.join(OUTPUT_PATH, f"{filename}.xlsx")
            wb.save(output_file)

            payload = {
                "success": True,
                "file_path": output_file,
                "entities_count": len(comparison_data)
            }
            return _tool_response(name, payload, started_at)

        except Exception as e:
            payload = {"error": str(e)}
            return _tool_response(name, payload, started_at)

    elif name == "create_time_series_report":
        filename = arguments["filename"]
        title = arguments["title"]
        company_name = arguments["company_name"]
        periods = arguments["periods"]
        metrics_by_period = arguments["metrics_by_period"]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Time Series"

            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(periods)+1)
            ws['A1'] = f"{title} - {company_name}"
            ws['A1'].font = Font(bold=True, size=16)

            # Get all unique metric names
            all_metrics = set()
            for period_metrics in metrics_by_period.values():
                for m in period_metrics:
                    all_metrics.add(m.get("metric_name", ""))

            # Headers: Metric | Period1 | Period2 | ...
            headers = ["Metric"] + periods
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = HEADER_FONT_WHITE
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER

            # Data rows
            row_idx = 4
            for metric_name in sorted(all_metrics):
                ws.cell(row=row_idx, column=1, value=metric_name).border = THIN_BORDER

                for col_idx, period in enumerate(periods, 2):
                    period_metrics = metrics_by_period.get(period, [])
                    value = None
                    for m in period_metrics:
                        if m.get("metric_name") == metric_name:
                            value = m.get("value")
                            break
                    ws.cell(row=row_idx, column=col_idx, value=value).border = THIN_BORDER

                row_idx += 1

            # Auto-adjust
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            output_file = os.path.join(OUTPUT_PATH, f"{filename}.xlsx")
            wb.save(output_file)

            payload = {
                "success": True,
                "file_path": output_file,
                "periods_count": len(periods),
                "metrics_count": len(all_metrics)
            }
            return _tool_response(name, payload, started_at)

        except Exception as e:
            payload = {"error": str(e)}
            return _tool_response(name, payload, started_at)

    else:
        payload = {"error": f"Unknown tool: {name}"}
        return _tool_response(name, payload, started_at)


async def main():
    async with stdio_server() as (read_stream, write_stream):
        await server.run(read_stream, write_stream, server.create_initialization_options())


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
